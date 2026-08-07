"""Relatório .xlsx dos leads enriquecidos.

Três abas em vez de uma planilha gigante, porque os dados têm cardinalidades
diferentes: empresa é 1 linha, mas cada empresa tem N contatos e N notícias.
Achatar tudo numa aba só obrigaria a repetir a empresa a cada notícia (ou a
espremer listas dentro de uma célula), que é justamente o que impede filtrar e
ordenar no Excel:

- **Empresas**  — uma linha por empresa, com todos os campos do enriquecimento;
- **Contatos**  — uma linha por decisor, ligada pelo CNPJ e pelo nome;
- **Notícias**  — uma linha por notícia, na mesma ligação.

Escolhi .xlsx e não .pdf de propósito: este relatório é insumo de trabalho
comercial (filtrar por porte, ordenar por ICP, colar num CRM), não documento
de leitura. PDF só faria sentido se fosse para apresentar/imprimir.
"""
import io
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from prospect_agent.services.enrichment_rules import formatar_telefone
from prospect_agent.services.normalizers import formatar_cnpj

# Identidade visual do produto (mesmo azul do BTN_GRADIENT).
_AZUL_CABECALHO = "1D548C"
_FONTE_CABECALHO = Font(bold=True, color="FFFFFF", size=11)
_FUNDO_CABECALHO = PatternFill("solid", fgColor=_AZUL_CABECALHO)

# (rótulo, largura). A largura é fixa porque autofit não existe no formato:
# o Excel calcula na abertura só se a coluna nunca foi dimensionada.
COLUNAS_EMPRESAS: Tuple[Tuple[str, int], ...] = (
    ("Empresa", 38), ("Razão social", 38), ("CNPJ", 20),
    ("Status cadastral", 16), ("Alerta de situação", 22), ("Alerta desde", 14),
    ("Cidade", 22), ("UF", 6), ("Bairro", 20), ("Endereço", 40), ("CEP", 12),
    ("Porte", 10), ("Faturamento estimado", 26), ("Segmento", 34),
    ("Início de atividade", 18), ("Idade (anos)", 12),
    ("Telefone", 18), ("WhatsApp", 10), ("Website", 32), ("LinkedIn", 40),
    ("Contatos", 10), ("Notícias", 10),
    ("% enriquecido", 14), ("Situação do enriquecimento", 22),
)

COLUNAS_CONTATOS: Tuple[Tuple[str, int], ...] = (
    ("Empresa", 38), ("CNPJ", 20), ("Contato", 30), ("Cargo", 34),
    ("Senioridade", 22), ("Área", 22), ("Origem", 12),
    # E-mail e confiança em colunas separadas de propósito: é planilha de
    # trabalho, e juntar os dois num texto só impediria filtrar por
    # "confiança >= 80" antes de importar num CRM.
    ("E-mail", 34), ("Confiança do e-mail (%)", 22),
    ("Perfil", 46),
)

COLUNAS_NOTICIAS: Tuple[Tuple[str, int], ...] = (
    ("Empresa", 38), ("CNPJ", 20), ("Data", 12), ("Título", 60),
    ("Resumo", 80), ("Fonte", 46),
)

_STATUS_PT = {
    "completed": "Completo",
    "partial": "Parcial",
    "failed": "Falhou",
    "pending": "Não enriquecido",
    "in_progress": "Em andamento",
}


def status_em_pt(status: Optional[str]) -> str:
    return _STATUS_PT.get(status or "", status or "")


def _escrever_cabecalho(aba, colunas: Tuple[Tuple[str, int], ...]) -> None:
    for i, (rotulo, largura) in enumerate(colunas, start=1):
        celula = aba.cell(row=1, column=i, value=rotulo)
        celula.font = _FONTE_CABECALHO
        celula.fill = _FUNDO_CABECALHO
        celula.alignment = Alignment(vertical="center")
        aba.column_dimensions[get_column_letter(i)].width = largura
    aba.freeze_panes = "A2"          # cabeçalho fixo ao rolar
    aba.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"


def montar_relatorio(
    empresas: List[Any],
    contatos_por_empresa: Dict[int, List[Any]],
    noticias_por_empresa: Dict[int, List[dict]],
    *,
    regiao: str = "",
    segmento: str = "",
) -> bytes:
    """Monta o .xlsx em memória e devolve os bytes (para rx.download)."""
    wb = Workbook()

    aba = wb.active
    aba.title = "Empresas"
    _escrever_cabecalho(aba, COLUNAS_EMPRESAS)
    for c in empresas:
        contatos = contatos_por_empresa.get(c.id, [])
        noticias = noticias_por_empresa.get(c.id, [])
        aba.append([
            c.nome,
            c.razao_social or "",
            formatar_cnpj(c.cnpj) or "",
            c.status_cadastral or "",
            c.alerta_situacao or "",
            c.alerta_situacao_desde or "",
            c.cidade or "",
            c.estado or "",
            c.bairro or "",
            c.endereco or "",
            c.cep or "",
            c.porte or "",
            c.faturamento_estimado or "",
            c.segmento or c.segmento_identificado or "",
            c.data_inicio_atividade or "",
            c.idade_empresa_anos if c.idade_empresa_anos is not None else "",
            # Formata na leitura, como a tabela: linhas gravadas antes da
            # padronização estão cruas no banco e sairiam "2121563600".
            formatar_telefone(c.telefone) or "",
            "Sim" if c.telefone_whatsapp else "",
            c.website_principal or c.website or "",
            c.linkedin_url or "",
            len(contatos),
            len(noticias),
            # Fração + formato de porcentagem: no Excel 91 vira "9100%".
            (c.enrichment_percentage or 0) / 100,
            status_em_pt(c.enrichment_status),
        ])

    # Índice pelo rótulo, não por aritmética sobre o tamanho da tupla: mexer nas
    # colunas não pode silenciosamente formatar a coluna errada.
    col_pct = get_column_letter(
        [r for r, _ in COLUNAS_EMPRESAS].index("% enriquecido") + 1
    )
    for linha in range(2, aba.max_row + 1):
        aba[f"{col_pct}{linha}"].number_format = "0%"

    aba_ct = wb.create_sheet("Contatos")
    _escrever_cabecalho(aba_ct, COLUNAS_CONTATOS)
    for c in empresas:
        for ct in contatos_por_empresa.get(c.id, []):
            aba_ct.append([
                c.razao_social or c.nome,
                formatar_cnpj(c.cnpj) or "",
                ct.nome,
                ct.cargo or "",
                ct.senioridade or "",
                ct.area or "",
                # "qsa" é jargão da Receita: no relatório vai o que significa.
                "Quadro societário" if ct.origem == "qsa" else "LinkedIn",
                ct.email or "",
                ct.email_confianca if ct.email else "",
                ct.perfil_url or "",
            ])

    aba_nt = wb.create_sheet("Notícias")
    _escrever_cabecalho(aba_nt, COLUNAS_NOTICIAS)
    for c in empresas:
        for n in noticias_por_empresa.get(c.id, []):
            aba_nt.append([
                c.razao_social or c.nome,
                formatar_cnpj(c.cnpj) or "",
                n.get("data_publicacao") or "",
                n.get("titulo") or "",
                n.get("resumo") or "",
                n.get("url") or n.get("fonte") or "",
            ])

    # Metadados: quem abrir o arquivo meses depois precisa saber de qual
    # pesquisa ele veio.
    wb.properties.title = "Relatório de enriquecimento da Plataforma Coester"
    if regiao or segmento:
        wb.properties.subject = f"{segmento} · {regiao}".strip(" ·")
    wb.properties.creator = "Plataforma Coester"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nome_do_arquivo(regiao: str, segmento: str, quando) -> str:
    """Nome previsível e ordenável: data primeiro, sem acento nem espaço."""
    import re
    import unicodedata

    def _slug(texto: str) -> str:
        sem_acento = "".join(
            ch for ch in unicodedata.normalize("NFD", texto or "")
            if unicodedata.category(ch) != "Mn"
        )
        limpo = re.sub(r"[^A-Za-z0-9]+", "-", sem_acento).strip("-").lower()
        return limpo[:40]

    partes = [quando.strftime("%Y-%m-%d"), "enriquecimento"]
    for t in (segmento, regiao):
        s = _slug(t)
        if s:
            partes.append(s)
    return "-".join(partes) + ".xlsx"
